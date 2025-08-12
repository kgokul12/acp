#!/usr/bin/env python3
import os
import subprocess
import signal
import sys
import re
import concurrent.futures

# Define temporary files
tmpfile = "/tmp/commits"
applied = "/tmp/applied"
sorted_file = "/tmp/sorted"
acp_log = "/home/amd/acp_log"
temp_file_path = "/tmp/tmp_file"
temp_script_path = "/tmp/tmp_script.sh"
acp_home = "/home/amd/.acp"

sorted_commits=[]
applied_commits=[]
Continue_flag = False

# Function to handle trapped signals
def Sig_catch(signum, frame):
    print("Processing Interrupt...")
    Run_command("git cherry-pick --abort")
    Reset_editor()
    print("cherry-pick aborted...")
    sys.exit(1)

def Trap_signals():
    # Trapping signal SIGINT and SIGTERM
    signal.signal(signal.SIGINT, Sig_catch)
    signal.signal(signal.SIGTERM, Sig_catch)

def Release_signals():
    # Untrap the signals by resetting them to default behavior
    signal.signal(signal.SIGINT, signal.SIG_DFL)
    signal.signal(signal.SIGTERM, signal.SIG_DFL)

def Change_core_editor():
    global temp_script_path
    # Create a temporary script to simulate an automated editor
    with open(temp_script_path,'w') as temp_script:
            temp_script.write("#!/bin/bash\nexit\n")
            temp_script_path = temp_script.name

    # Make the temporary script executable
    os.chmod(temp_script_path, 0o755)

    Run_command(f"git config core.editor {temp_script_path}")

def Reset_editor():
    #Reset editor back to "vim"
    if os.path.exists("$PWD/.git/.COMMIT_EDITMSG.swp"):
        os.remove("$PWD/.git/.COMMIT_EDITMSG.swp")
    Run_command(f"git config core.editor 'vi'")
    if os.path.exists(temp_file_path):
        os.remove(temp_file_path)
    os.remove(temp_script_path)

#add commit upstream message.
def Add_upstream_msg(commit):
    global temp_file_path

    # Modify the commit message to include the SHA_ID and additional text
    commit_message = subprocess.check_output("git log -1 --pretty=%B", shell=True, universal_newlines=True).strip()

    # Split the commit message into the first line and the rest
    lines = commit_message.split('\n', 1)
    first_line = lines[0]
    rest_of_message = lines[1] if len(lines) > 1 else ""

    # Combine the new messageg
    new_commit_message = f"{first_line}\n\ncommit {commit} upstream\n\n{rest_of_message}"

    # Write the new commit message to a temporary file
    with open(temp_file_path, mode='w') as temp_file:
        temp_file.write(new_commit_message)
        temp_file_path = temp_file.name

    # Amend the commit with the new message
    print("Adding upstream commit message...")
    Run_command(f"git commit --amend --file={temp_file_path}")

def Get_commit_input():
    commits=[]
    # Read multiple lines of commit IDs into the 'commits' list
    print("Enter the commit IDs (one per line). Press Ctrl+D when done:")
    try:
        while True:
            line = input()
            if line:
                commits.append(line)
    except EOFError:
        pass

    if os.path.exists(tmpfile):
        for commit in commits:
            Run_command(f"echo \"grep -n {commit} {acp_log}\" >> {tmpfile}")
    else:
        with open(tmpfile, "w") as f:
            for commit in commits:
                f.write(f"grep -n {commit} {acp_log}\n")

def Process_commits():
    global sorted_commits, applied_commits
    # Sort and unique the commit IDs, then prepare the grep commands
    if not os.path.exists(acp_log):
        print(f"\nMake a log copy of kernel latest version in Home as {acp_log}")
        print("    eg : checkout to kernel master branch ")
        print(f"         git log --pretty=oneline > {acp_log}")
        print(f"   (or) run acp log (or) acp -lg to create log file")
        sys.exit(1)

    os.chmod(tmpfile, 0o755)

    grep_output = subprocess.run(f"{tmpfile} | sort -ugr", stdout=subprocess.PIPE, universal_newlines=True, shell=True)
    if not grep_output.stdout.strip():
        print("No commits to process.. May be a invalid commit id")
        sys.exit(1)

    commits = grep_output.stdout.splitlines()

    # Save the output to a file
    with open(sorted_file, "w") as output_file:
        for commit in commits:
            output_file.write(f"{commit}\n")

    for line in commits:
            sorted_commits.append(line.split(':')[1].split()[0])
    print("Commits sorted....")

    # Take all the applied commits from the applied file in applied_commits variable
    if os.path.exists(applied):
        cat_output = subprocess.run(["cat", applied], stdout=subprocess.PIPE, universal_newlines=True)
        applied_commits = cat_output.stdout.splitlines()

# Definition to check the commit is already applied ot not
def Check_commit_status(commit, count):
    if commit in applied_commits and commit == applied_commits[count]:
        # if commit already applied in order skip and return -1
        return -1
    else :
        # Return the value to reset...
        return len(applied_commits) - count

def Apply_commits():
    # Applying the commits
    count = 0
    check_val = -1
    for commit in sorted_commits:
        if check_val == -1:
            check_val = Check_commit_status(commit, count)
            if check_val == -1:
                count += 1
                continue
            else:
                print(f"git reset --hard HEAD~{check_val}")
                Run_command(f"git reset --hard HEAD~{check_val}")
                check_val  = 0

        print(f" --> Applying commit {commit} ....")
        result = subprocess.run(["git", "cherry-pick", commit])
        print(result.returncode)
        if result.returncode != 0:
            print(f"\nConflict occurred while applying commit {commit}")
            print("Resolve the conflict and enter \n'done' / 'd' --> continue \n'abort' / 'a' --> cancel \n'bash' / 'b' --> open bash\n")

            try:
                while True:
                    user_input = input().strip().lower()
                    if user_input == "done" or user_input == 'd':
                        Run_command("git add -u")
                        Run_command("git cherry-pick --continue")
                        break
                    elif user_input == "abort" or user_input == 'a':
                        if os.path.exists("$PWD/.git/.COMMIT_EDITMSG.swp"):
                            os.remove("$PWD/.git/.COMMIT_EDITMSG.swp")
                        Reset_editor()
                        Run_command(f"git config core.editor 'vi'")
                        Run_command("git cherry-pick --abort")
                        sys.exit(1)
                    elif user_input == "bash" or user_input == 'b':
                        print("Entering bash mode... to come out enter 'exit'")
                        subprocess.run("bash")
                    else :
                        print("Invalid input...")
            except EOFError:
                break
        print(f"Commit {commit} successfully applied....")
        Add_upstream_msg(commit)
        Run_command(f"echo {commit} >> {applied}")

    # Display a completion message
    print("All commits have been processed.")

def Cleanup():
    # Clean up the temporary files
    if os.path.exists(tmpfile):
        os.remove(tmpfile)
    if os.path.exists(sorted_file):
        os.remove(sorted_file)
    if os.path.exists(applied):
        os.remove(applied)
    if os.path.exists(temp_file_path):
        os.remove(temp_file_path)
    if os.path.exists(temp_script_path):
        os.remove(temp_script_path)
    i=1
    while True:
        if os.path.exists(f"./{i}b"):
            os.remove(f"./{i}b")
        else :
            break
        if os.path.exists(f"./{i}u"):
            os.remove(f"./{i}u")
        else :
            break
        i+=1

def Auto_cherry_pick():
    if Continue_flag:
        if not os.path.exists(tmpfile):
            print("Add input commits first...")
            #get input commits, process, sort and save into sorted_commits...
            Get_commit_input()
        else :
            Process_commits()
    else :
        #get input commits, process, sort and save into sorted_commits...
        Get_commit_input()
        Process_commits()

    Trap_signals()

    Change_core_editor()

    Apply_commits()

    Reset_editor()

    Release_signals()

    Cleanup()



#===UTILS===
 
def Run_command(command):
    """Run a shell command and handle errors."""
    result = subprocess.run(command, stderr=subprocess.PIPE, shell=True, universal_newlines=True)
    if result.returncode != 0:
        print(result.stderr)
        print(f"Error: Command failed: {command}")
        exit(1)

#===DIFF===
def Check_commit_diff():
    # Local function Process diff
    def Process_diff(bp_commit,file_no):
        # check bp_commit
        result = subprocess.run(f"git show {bp_commit} > {file_no}b",stdout=subprocess.PIPE, shell=True, universal_newlines=True)
        grep_out = subprocess.run(f"grep upstream {file_no}b", stdout=subprocess.PIPE, shell=True, universal_newlines=True)
        if not grep_out.stdout:
            print(f"==> For {bp_commit} --> no Upstream id is there")
            print(f"Check commit message of {bp_commit} or check the repo")
            return 1

        up_commit = grep_out.stdout.strip().split()[1]
        result = subprocess.run(f"git show {up_commit}", stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, universal_newlines=True)
        if not result.stdout:
            print(f"==> For {bp_commit} --> no Upstream id is there")
            print(f"Check commit message of {bp_commit} or check the repo")
            return 1

        with open (f"{file_no}u",'w') as file:
            file.write(result.stdout)
        print(f"==> Compare {bp_commit} | {up_commit[:14]} :: diff {file_no}b {file_no}u")
        
        result = subprocess.run(f"diff {file_no}b {file_no}u | grep -e '> +' -e '> -' -e '< +' -e '< -'", stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, universal_newlines=True)
        diff = result.stdout.strip() or "!* no diff"
        if diff != "!* no diff":
            result = subprocess.run(f"grep -e 'Backport changes' -e 'Backport Changes' {file_no}b", stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, universal_newlines=True)
            print(f"Deviaton explaination : {'Yes' if result.stdout else 'No'}")
        print(diff)
    # Local function Process_diff ends here

    if len(sys.argv) < 3:
        print("Add commit id you want to check diff or give count to check the list of diffs")
        sys.exit(1)
    try:
        value = int(sys.argv[2])
        print("Checking diff for the following commits")
        result=subprocess.run(f"git log --oneline -{value}", stdout=subprocess.PIPE, shell=True, universal_newlines=True)
        print(result.stdout)
        logs = result.stdout.splitlines()
        bp_commits = []
        for log in logs:
            bp_commits.append(log.split()[0])
        file_no=1
        for bp_commit in bp_commits[::-1]:
            if Process_diff(bp_commit,file_no):
                continue
            file_no += 1

    except ValueError:
        # if commit is given as input
       Process_diff(sys.argv[2],1)    

#===UPDATE===
def Update_acp():
    if not os.path.exists(f"{acp_home}"):
        Run_command(f"git clone https://github.com/kgokul12/acp {acp_home}")
    else :
        Run_command(f"git -C {acp_home} pull origin main -q --rebase") 
    
    Run_command(f"make -C {acp_home}")
    Run_command("acp -h")
    sys.exit(0)

#===REVIEW SCRIPT===
try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    Run_command("sudo pip3 install openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def Create_review_links():
    if len(sys.argv) < 3:
        print("Usage: acp review <commit_count>")
        sys.exit(1)

    count = int(sys.argv[2])
    outfile = "/home/amd/commits.xlsx"

    result = subprocess.run(
        ["git", "log", "--pretty=oneline", f"-n{count}"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        universal_newlines=True,
        check=True
    )

    print(result.stdout)

    bpcmids = []
    names = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Backport approval sheet"

    # Header row
    headers = ["Subject", "Backport Commit", "Upstream Commit", "Approved"]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Apply header styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Parse commits
    list_commits = result.stdout.strip().split('\n')
    for commit in list_commits:
        commit_parts = commit.strip().split(' ')
        bpcmids.append(commit_parts[0])
        names.append(' '.join(commit_parts[1:]))

    for bpcmid, name in zip(bpcmids[::-1], names[::-1]):
        bplink = "https://github.com/AMDEPYC/Linux_Backport/commit/"
        uplink = "https://github.com/torvalds/linux/commit/"
        upcmid = ""

        try:
            result = subprocess.run(
                ["git", "show", f"{bpcmid}"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=True,
                check=True
            )
            grep_result = subprocess.run(
                ["grep", "upstream"],
                input=result.stdout,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=True,
                check=True
            )
            upcmid = grep_result.stdout.strip().split(' ')[1]
        except subprocess.CalledProcessError as e:
            print(f"Error finding upstream commit in {bpcmid}: {e.stderr}")
            upcmid = "N/A"

        # Write row
        bp = f'=HYPERLINK("{bplink + bpcmid}", "{bpcmid[:14]}")'
        up = f'=HYPERLINK("{uplink + upcmid}", "{upcmid[:14]}")' if upcmid != "N/A" else "N/A"
        row = [name, bp, up]

        current_row = ws.max_row + 1
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Set fixed column widths
    column_widths = [50, 22, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # format hyperlink
    hyperlink_font = Font(color="467886", underline="single")
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK('):
                cell.font = hyperlink_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Saving file
    wb.save(outfile)
    print(f"xlsx file created at {outfile}")
    
#===COMMIT AVAILABILITY SCRIPT===
 
def get_current_branch():
    return subprocess.check_output(["git", "branch", "--show-current"]).strip().decode()

def get_commit_details(cmid):
    return subprocess.check_output(["git", "show", cmid]).strip().decode().splitlines()[0]

def Check_avail():
    # Check if any arguments are passed
    if len(sys.argv) == 2:
        print("To skip copying log, use\n\t$ ./commit.py -i\n\t*only if you copied the log file in /tmp/log\n")
    
    print(f"\nChecking commits in branch --> {get_current_branch()}\n")
    print("Enter all the commits")

    if len(sys.argv) == 2:
        print("Wait, It's logging....")
        Run_command(f"git log > /tmp/log")
        Run_command(f"git log --pretty=oneline > /tmp/1_log")

    print("\n ######################################### \n")

    x = 0  # Available commits count
    y = 0  # Not Available commits count
    z = 0  # May be as Upstream

    # Read multiple lines of commit IDs into the 'commits' list
    commits = []
    print("Now, Enter the commit IDs (one per line). Press Ctrl+D when done:")
    
    try:
        while True:
            line = input()
            if line :
                 commits.append(line)
    except EOFError:
        pass
    if os.path.exists("/tmp/Avail_commit"):
        os.remove("/tmp/Avail_commit")
    if os.path.exists("/tmp/No_Avail_commit"):
        os.remove("/tmp/No_Avail_commit")

    Run_command("touch /tmp/Avail_commit /tmp/No_Avail_commit")
    for cmid in commits:
        res = subprocess.run(f"grep {cmid} /tmp/1_log -q", shell=True, universal_newlines=True)
        if not res.returncode :
            x += 1
            Run_command(f"echo {cmid} >> /tmp/Avail_commit")
            print(f"{get_commit_details(cmid)}   Available")
        else :
            grep_line = subprocess.run(f"grep {cmid} /tmp/log", stderr=subprocess.PIPE, stdout=subprocess.PIPE, shell=True, universal_newlines=True)
            if grep_line.stdout:
                z += 1
                Run_command(f"echo {cmid} >> /tmp/Avail_commit")
                print(f"commit {cmid} Available like this ")
                for line in grep_line.stdout.splitlines():
                    print(f"\t{line}")
            else:
                y += 1
                Run_command(f"echo {cmid} >> /tmp/No_Avail_commit")
                print(f"commit {cmid}   Not Available")

    print("\n ######################################### \n")
    print(f"\tTotally   -({x})- commits Available directly -({z})- may be as upstream and -({y})- commits Not Available\n")
    print("\tAvailable commits are stored in --> /tmp/Avail_commit")
    print("\t   and not available commits in --> /tmp/No_Avail_commit")

#===CHECK DEPS===

def Check_deps():
    # Define a function to check each commit in parallel
    def check_commit(cmid):
        short_cmid = cmid[:12]

        # Check for dependencies using `--grep`
        deps_result = subprocess.run(f"git log origin/stable_kernel_master --oneline --grep=\"{short_cmid}\"", stdout=subprocess.PIPE, shell=True, universal_newlines=True)
        deps_output = deps_result.stdout.strip() or "no deps"
        return f"==>checking {short_cmid}\n{deps_output}"

    # function to process commit ids if not given as input
    def Process_upstream(up_cmids,count):
        log_out = subprocess.run(f"git log -{count} ", stdout=subprocess.PIPE, shell=True, universal_newlines=True)
        lines = [line for line in log_out.stdout.splitlines()]

        pattern = rf"\b{re.escape('commit')}\s+(\w+)\s+{re.escape('upstream')}\b"

        for line in lines:
            if re.search(pattern, line):
                up_cmids.append(line.strip().split()[1])

    #Function to process commit ids given as input
    def getlist(input_list):
        print("Enter a value (or type 'done' to finish): ")
        while True:
            user_input = input()
            if user_input.lower() == 'done':
                 break
            input_list.append(user_input)

    #--MAIN--
    # check the arguments
    if len(sys.argv) < 3:
        print("acp <ck-dep/-cdp> <count/-l>")
        print("Enter the count(inside Linux_Backport repo) or -l(to add a list of commit ids) to check the deps")
        sys.exit(0)

    up_cmids = []
    count=0

    try:
        ret_val = int(sys.argv[2])
        count = sys.argv[2]
    except ValueError:
        if sys.argv[2] == "-l":
            getlist(up_cmids)
        else :
            print("Invalid input")
            sys.exit(1)

    if count != 0 :
        Process_upstream(up_cmids,count)

    print(up_cmids)
    # Use a ThreadPoolExecutor for parallel processing
    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = list(executor.map(check_commit, up_cmids))

    # Print all results
    for result in results:
        print(result)

#===ACP MAIN===
 
def acp_help():
    print("\033[95mOptions related to cherry-pick\033[0m")
    print("     acp [options] (-l, -c, -r, -a)")
    print("     acp [options] <count/commit> (-s)")
    print("         -a or add      --> to add a commit id to existing list, to create a new list use -c first and then use -a")
    print("         -l or list     --> to check the list of ordered commit ids")
    print("         -c or continue --> to clear of logs of cherry-pick")
    print("         -s or status   --> to check the status of applied commits and list of commits in order")
    print("\n\033[95mOptions for complex git operations\033[0m")
    print("         -r or reset    --> to reset the logs")
    print("                        --> -r all or --reset all to reset all acp applied commits")
    print("         -S or Signoff  --> to add signoff message to the commit log \"acp -S <count/commit_id>\"")
    print("\n\033[95mOptions for review and others\033[0m")
    print("         -d or diff     --> to check the diff of a backported patch")
    print("         -cal or ck-avl  --> to check the availability of the commits 'Usage : acp ck-avl <options>'")
    print("         -cdp or ck-dep  --> to check the availability of the commits 'Usage : acp ck-dep <options>'")
    print("         -cf or ck-files --> to check the files backported are compiled or not 'Usage : acp ck-file'")
    print("         order          --> to order you backport commit ids as per upstream")
    print("         review         --> to create review request links 'Usage : acp review <count>'")
    print("\n\033[95mOthers\033[0m")
    print("         sysinfo        --> to get the Host system info used for backporting")
    print("         mk-link        --> to create link")
    print("         ck-tag         --> to create link")
    print("         -u or update   --> to update your script")
    print("         -lg or log     --> to update your script")
    print("         -cl or clean   --> to clear of logs of cherry-pick")
    print(f"\n\033[93mSources available at {acp_home}\033[0m")

#===OPTIONS===
def Call_options():
    global Continue_flag

    if sys.argv[1] in ["reset", "-r"]:
        if sys.argv[1] in ["-r", "reset"] and len(sys.argv) > 2 and sys.argv[2] in ["all", "a"]:
            if os.path.exists(applied):
                with open(applied, 'r') as file:
                    count = sum(1 for line in file)
                print(f"==> git reset --hard HEAD~{count}")
                Run_command(f"git reset --hard HEAD~{count}")
                print("Cleaning all applied commits")
                os.remove(applied)
            sys.exit(0)
        else:
            try:
                # Try to convert sys.argv[2] to an integer
                reset_value = int(sys.argv[2])
                print(f"==> git reset --hard HEAD~{reset_value}")
                Run_command(f"git reset --hard HEAD~{reset_value}")
            except ValueError:
                # If conversion fails, it's not an integer, so treat it as a commit hash
                print(f"==> git reset --hard {sys.argv[2]}")
                Run_command(f"git reset --hard {sys.argv[2]}")
            sys.exit(0)
            # change in applied_file pending....

    elif sys.argv[1] in ["status", "-s"]:
        if not os.path.exists(sorted_file):
            print("Nothing in list to show")
            sys.exit(0)
        print("\nThe list of commits in QUEUE are:\n")
        Run_command(f"cat {sorted_file}")
        if not os.path.exists(applied):
            print("\nNo commits applied yet\n")
            sys.exit(0)
        print("\nThe list of commits already applied:\n")
        Run_command(f"cat {applied}")
        sys.exit(0)

    elif sys.argv[1] in ["list", "-l"]:
        if not os.path.exists(sorted_file) and not os.path.exists(tmpfile):
            print("Nothing added to list")
            sys.exit(1)
        else :
            Process_commits()
            print("The list of commits in QUEUE are:")
            Run_command(f"cat {sorted_file}")
            sys.exit(0)

    elif sys.argv[1] in ["clean", "-cl"]:
        Cleanup()
        print("Cleaning done successfully..")
        sys.exit(0)

    elif sys.argv[1] in ["add", "-a"]:
        Get_commit_input()
        print("Added successfully to list....")
        sys.exit(0)
        
    elif sys.argv[1] in ["diff", "-d"]:
        Check_commit_diff()
        sys.exit(0)

    elif sys.argv[1] in ["ck-avl", "-cal"]:
        Check_avail()
        sys.exit(0)

    elif sys.argv[1] in ["ck-dep", "-cdp"]:
        Check_deps()
        sys.exit(0)

    elif sys.argv[1] in ["ck-files", "-cf"]:
        Run_command(f"{acp_home}/file_check.py")
        sys.exit(0)

    elif sys.argv[1]=="review":
        Create_review_links()
        sys.exit(0)
        
    elif sys.argv[1]=="sysinfo":
        Run_command(f"{acp_home}/sysinfo.sh")
        sys.exit(0)
        
    elif sys.argv[1]=="order":
        Run_command(f"{acp_home}/order.sh")
        sys.exit(0)
    elif sys.argv[1]=="mk-link":
        Run_command(f"{acp_home}/hyperlink_maker.py")
        sys.exit(0)
    elif sys.argv[1]=="ck-tag":
        Run_command(f"{acp_home}/find_tag.py")
        sys.exit(0)

    elif sys.argv[1] in ["signoff", "-S"]:
        # Run `git commit --amend -s`
        print("Adding sign-off content...")
        try:
            # Try to convert sys.argv[2] to an integer
            reset_value = int(sys.argv[2])
            print(f"==> git rebase --signoff HEAD~{reset_value}")
            Run_command(f"git rebase --signoff HEAD~{reset_value}")
        except ValueError:
            # If conversion fails, it's not an integer, so treat it as a commit hash
            print(f"==> git rebase --signoff {sys.argv[2]}")
            Run_command(f"git rebase --signoff {sys.argv[2]}")
        sys.exit(0)

    elif sys.argv[1] in ["continue", "-c"]:
        Continue_flag = True
        
    elif sys.argv[1] in ["update", "-u"]:
        Update_acp()
        sys.exit(0)
    
    elif sys.argv[1] in ["log", "-lg"]:
        repo = input("Enter the repo link you want to take log from : ")
        branch = input("Enter the name of the branch : ")
        dir = "/tmp/tmp_repo"
        if os.path.exists(dir):
            Run_command(f"rm -rf {dir}")
        print(f"git clone --single-branch -b {branch} {repo} {dir} -j10")
        print("!Downloading repo.....")
        Run_command(f"git clone --single-branch -b {branch} {repo} {dir} -j10")
        print("!Updating log.....")
        Run_command(f"git -C {dir} log --pretty=oneline {branch} > {acp_log}")
        print("log updated successfully to {acp_log}")
        sys.exit(0)

    elif sys.argv[1] == "--help" or sys.argv[1] == "-h":
        acp_help()
        sys.exit(0)

    else:
        print("Invalid option... See help")
        sys.exit(1)

#===MAIN===
if __name__ == "__main__":
    if len(sys.argv) > 1:
        Call_options()
    
    Auto_cherry_pick()
